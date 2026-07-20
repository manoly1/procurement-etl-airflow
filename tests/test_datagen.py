"""Tests for the synthetic data generator."""

from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import load_workbook

from datagen.cli import parse_weeks
from datagen.datasets import DATASETS
from datagen.generate import generate_file, generate_frame, report_monday


def test_report_monday_is_iso_week_monday() -> None:
    # ISO week 30 of 2026 starts on Monday 2026-07-20.
    assert report_monday(30).isoformat() == "2026-07-20"


@pytest.mark.parametrize("name", list(DATASETS))
def test_clean_frame_has_expected_columns(name: str) -> None:
    df, flags = generate_frame(name, 29, dirty_on=False)
    assert flags == set()
    assert list(df.columns) == list(DATASETS[name].columns)
    assert len(df) == 60


def test_generation_is_deterministic_with_seed() -> None:
    a, _ = generate_frame("open_po", 29, dirty_on=True, seed=123)
    b, _ = generate_frame("open_po", 29, dirty_on=True, seed=123)
    pd.testing.assert_frame_equal(a, b)


def test_leading_zeros_archetype() -> None:
    df, _ = generate_frame("open_po", 29, dirty_on=True, seed=1)
    mats = df["Material"].dropna().astype(str)
    assert mats.str.match(r"^0*\d{8}$").any()
    assert mats.str.startswith("0").any()


def test_numbers_as_text_archetype() -> None:
    df, _ = generate_frame("open_po", 29, dirty_on=True, seed=1)
    values = df["Net Price"].tolist()
    assert any(isinstance(v, str) and "," in v for v in values)


def test_missing_keys_archetype() -> None:
    ds = DATASETS["open_po"]
    df, _ = generate_frame("open_po", 29, dirty_on=True, seed=1)
    assert df[ds.key_cols[-1]].isna().any()


def test_clean_frame_has_no_dirt() -> None:
    df, _ = generate_frame("open_po", 29, dirty_on=False, seed=1)
    assert not df["PO Item"].isna().any()
    assert pd.api.types.is_integer_dtype(df["Material"])  # ints, zeros not forced


def test_dirty_file_has_title_row_then_headers(tmp_path) -> None:
    path = generate_file("open_po", 29, dirty_on=True, out_dir=str(tmp_path))
    assert path.endswith("open_po/week=29/open_po_W29.xlsx")
    ws = load_workbook(path).active
    assert str(ws.cell(1, 1).value).startswith("Report Date")
    assert ws.cell(2, 1).value == "PO Number"


def test_clean_file_has_headers_on_row_one(tmp_path) -> None:
    path = generate_file("all_prs", 25, dirty_on=False, out_dir=str(tmp_path))
    ws = load_workbook(path).active
    assert ws.cell(1, 1).value == "PR Number"


def test_unknown_dataset_raises() -> None:
    with pytest.raises(ValueError):
        generate_frame("nope", 29)


@pytest.mark.parametrize(
    "spec,expected",
    [("29", [29]), ("25-27", [25, 26, 27]), ("25,27,29", [25, 27, 29])],
)
def test_parse_weeks(spec: str, expected: list[int]) -> None:
    assert parse_weeks(spec) == expected
